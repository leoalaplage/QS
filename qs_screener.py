#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
QS SCREENER v3  -  Screener d'actions "Quality Stocks"
======================================================

Reproduit et prolonge le systeme de notation du Google Sheet "QS Screener v2".

Chaque action recoit un rang-percentile 0-100 sur chaque metrique, calcule
DANS l'univers fourni (auto-calibre). Ces percentiles sont agreges en 4 piliers
(Quality / Health / Growth / Value) puis en un score TOTAL, avec classement et
compteur d'alertes.

Ajouts v3 (au-dela de la feuille) :
  * note-lettre (A+ ... D) sur le score TOTAL ;
  * score de CONVICTION = TOTAL ajuste du risque (malus par alerte) ;
  * comparaison INTRA-SECTEUR (percentiles recalcules dans le secteur) ;
  * FORCES / FAIBLESSES automatiques par action ;
  * classeur Excel avec un onglet Synthese (tableau de bord) + Fiches.

Usage minimal :
    python3 qs_screener.py data/univers_exemple.csv

Tous les reglages sont dans qs_config.py. Voir README.md.
"""

import argparse
import csv
import os
import subprocess
import sys
import unicodedata

import qs_config as cfg
import qs_pdf

PILIERS = ["Quality", "Health", "Growth", "Value"]


# ===========================================================================
#  Utilitaires
# ===========================================================================
def _norm(texte):
    if texte is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texte))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return "".join(c for c in t.lower() if c.isalnum())


# Suffixes d'ordre de grandeur, ramenes en "milliards" (base des cap. boursieres).
_MULT = {"K": 1e-6, "M": 1e-3, "B": 1.0, "T": 1e3}


def _to_float(valeur):
    """Convertit une cellule en float. Robuste aux exports type fiscal.ai :
    symboles monetaires, %, separateurs de milliers, suffixes B/M/T/K,
    negatifs entre parentheses ou avec tiret. Renvoie None si vide."""
    if valeur is None:
        return None
    s = str(valeur).strip()
    if s == "" or s.upper() in ("N/A", "NA", "#N/A", "NAN", "NEUTRAL", "-", "--", "NM", "NMF"):
        return None

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1]
    for ch in (" ", " ", " ", "$", "€", "£", "¥", "%"):
        s = s.replace(ch, "")
    if s[:1] == "+":
        s = s[1:]
    if s[:1] in ("-", "−"):
        neg, s = True, s[1:]

    mult = 1.0
    if s and s[-1] in "KkMmBbTt" and any(c.isdigit() for c in s[:-1]):
        mult, s = _MULT[s[-1].upper()], s[:-1]

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        v = float(s) * mult
        return -v if neg else v
    except ValueError:
        return None


def _par_action(cagr_total, cagr_actions):
    """Croissance par action : (1+g_total)/(1+g_actions) - 1, en %.
    cagr_actions = CAGR du nombre d'actions (positif = dilution)."""
    if cagr_total is None or cagr_actions is None:
        return None
    denom = 1.0 + cagr_actions / 100.0
    if denom <= 0:
        return None
    return ((1.0 + cagr_total / 100.0) / denom - 1.0) * 100.0


def _score_absolu(cle, valeur):
    """Score absolu 0-100 d'une valeur brute, d'apres les ancres de config.
    Renvoie None si pas d'ancre / valeur manquante."""
    if valeur is None:
        return None
    ancre = cfg.ANCRES_ABSOLUES.get(cle)
    if ancre is None:
        return None
    if cle in cfg.NEGATIF_PIRE and valeur <= 0:   # multiple negatif = absurde
        return 0.0
    v0, v100 = ancre
    if v100 == v0:
        return 50.0
    score = (valeur - v0) / (v100 - v0) * 100.0
    return max(0.0, min(100.0, score))


def _percentile(valeurs_triees, p):
    n = len(valeurs_triees)
    if n == 0:
        return None
    if n == 1:
        return valeurs_triees[0]
    rang = (p / 100.0) * (n - 1)
    bas = int(rang)
    frac = rang - bas
    if bas + 1 < n:
        return valeurs_triees[bas] + frac * (valeurs_triees[bas + 1] - valeurs_triees[bas])
    return valeurs_triees[bas]


def _mediane(valeurs):
    v = sorted(x for x in valeurs if x is not None)
    n = len(v)
    if n == 0:
        return None
    m = n // 2
    return v[m] if n % 2 else (v[m - 1] + v[m]) / 2.0


def _rang_competition(indices_ordonnes, cle_valeur):
    """Renvoie {index: rang} en classement de competition (ex aequo = meme rang)."""
    rangs, rang, precedent = {}, 0, None
    for pos, i in enumerate(indices_ordonnes, start=1):
        v = round(cle_valeur(i), 4)
        if v != precedent:
            rang, precedent = pos, v
        rangs[i] = rang
    return rangs


# ===========================================================================
#  Chargement du CSV et mapping des colonnes
# ===========================================================================
def _trouver_colonne(entetes_norm, alias_list):
    for alias in alias_list:
        cle = _norm(alias)
        if cle in entetes_norm:
            return entetes_norm[cle]
    return None


def _charger_map_secteurs():
    """Charge secteurs.csv (Ticker,Secteur) situe a cote du script. Sert a
    completer le secteur quand le CSV d'entree n'en fournit pas."""
    chemin = os.path.join(os.path.dirname(os.path.abspath(__file__)), "secteurs.csv")
    mapping = {}
    if not os.path.exists(chemin):
        return mapping
    try:
        with open(chemin, "r", encoding="utf-8-sig", newline="") as f:
            echantillon = f.read(2048)
            f.seek(0)
            try:
                delim = csv.Sniffer().sniff(echantillon, delimiters=",;\t").delimiter
            except csv.Error:
                delim = ","
            for ligne in csv.DictReader(f, delimiter=delim):
                cles = {_norm(k): k for k in (ligne.keys() or [])}
                kt, ks = cles.get(_norm("Ticker")), cles.get(_norm("Secteur"))
                if kt and ks:
                    tk = (ligne[kt] or "").strip()
                    sc = (ligne[ks] or "").strip()
                    if tk and sc:
                        mapping[_norm(tk)] = sc
    except (OSError, csv.Error):
        pass
    return mapping


def charger_csv(chemin, delimiteur=None):
    with open(chemin, "r", encoding="utf-8-sig", newline="") as f:
        echantillon = f.read(4096)
        f.seek(0)
        if delimiteur is None:
            try:
                delimiteur = csv.Sniffer().sniff(echantillon, delimiters=",;\t").delimiter
            except csv.Error:
                delimiteur = ","
        lecteur = csv.DictReader(f, delimiter=delimiteur)
        entetes = lecteur.fieldnames or []
        lignes = list(lecteur)

    entetes_norm = {_norm(h): h for h in entetes}

    col_ticker = _trouver_colonne(entetes_norm, cfg.COLONNE_TICKER)
    if col_ticker is None:
        raise SystemExit("ERREUR : aucune colonne 'Ticker' trouvee dans le CSV.")
    col_secteur = _trouver_colonne(entetes_norm, cfg.COLONNE_SECTEUR)
    col_cap = _trouver_colonne(entetes_norm, cfg.COLONNE_CAP)

    mapping = {}
    for m in cfg.METRIQUES:
        mapping[m["cle"]] = _trouver_colonne(entetes_norm, m["entetes"])
    mapping_ref = {r["cle"]: _trouver_colonne(entetes_norm, r["entetes"])
                   for r in cfg.COLONNES_REFERENCE}
    map_secteurs = _charger_map_secteurs()

    titres = []
    for ln in lignes:
        ticker = (ln.get(col_ticker) or "").strip()
        if ticker == "":
            continue
        secteur = (ln.get(col_secteur) or "").strip() if col_secteur else ""
        if not secteur:
            secteur = map_secteurs.get(_norm(ticker), "(n/a)")
        rec = {
            "Ticker": ticker,
            "Secteur": secteur,
            "Cap": _to_float(ln.get(col_cap)) if col_cap else None,
            "brut": {}, "ref": {},
        }
        for m in cfg.METRIQUES:
            col = mapping[m["cle"]]
            rec["brut"][m["cle"]] = _to_float(ln.get(col)) if col else None
        for r in cfg.COLONNES_REFERENCE:
            col = mapping_ref[r["cle"]]
            rec["ref"][r["cle"]] = _to_float(ln.get(col)) if col else None
        # OCF/Capex derive si la colonne ratio est absente mais OCF & Capex presents
        if rec["brut"].get("OCF_Capex") is None:
            ocf, capex = rec["ref"].get("OCF"), rec["ref"].get("Capex")
            if ocf is not None and capex not in (None, 0):
                rec["brut"]["OCF_Capex"] = ocf / abs(capex)
        # Croissance PAR ACTION = CAGR de la metrique corrigee du CAGR du nb d'actions
        dil = rec["brut"].get("ShOut5")
        if rec["brut"].get("RevPS5") is None:
            rec["brut"]["RevPS5"] = _par_action(rec["brut"].get("Rev5"), dil)
        if rec["brut"].get("FCFPS5") is None:
            rec["brut"]["FCFPS5"] = _par_action(rec["brut"].get("LevFCF5"), dil)
        titres.append(rec)

    if not titres:
        raise SystemExit("ERREUR : aucune ligne de donnees exploitable dans le CSV.")

    # metrique reellement "manquante" = aucune valeur exploitable sur tout l'univers
    manquantes = [m["cle"] for m in cfg.METRIQUES
                  if all(t["brut"].get(m["cle"]) is None for t in titres)]
    return titres, manquantes


# ===========================================================================
#  Coeur du scoring
# ===========================================================================
def _valeur_utilisable(cle, v):
    """None si donnee absente ; multiples de valorisation <= 0 = absurdes -> exclus
    du classement (traites comme 'pire' ailleurs)."""
    if v is None:
        return None
    if cle in cfg.NEGATIF_PIRE and v <= 0:
        return None
    return v


def _percentiles_groupe(groupe, cle_cible, winsoriser):
    """Percentiles par metrique DANS `groupe` -> t[cle_cible][metrique].
    Les valeurs manquantes / absurdes ne recoivent PAS 50 : la cle est laissee
    absente et sera geree (renormalisation / pire) au niveau du score."""
    for t in groupe:
        t[cle_cible] = {}
    for m in cfg.METRIQUES:
        cle, sens = m["cle"], m["sens"]
        presents = [(t, _valeur_utilisable(cle, t["brut"][cle])) for t in groupe]
        presents = [(t, v) for t, v in presents if v is not None]
        valeurs = [v for _, v in presents]

        # plafond economique : au-dela, "plus haut" n'est plus "meilleur"
        plafond = cfg.PLAFONDS.get(cle) if hasattr(cfg, "PLAFONDS") else None
        if plafond is not None:
            valeurs = [min(v, plafond) for v in valeurs]

        if winsoriser and cfg.WINSOR_BAS is not None and len(valeurs) >= 3:
            tri = sorted(valeurs)
            lo, hi = _percentile(tri, cfg.WINSOR_BAS), _percentile(tri, cfg.WINSOR_HAUT)
            valeurs = [min(max(v, lo), hi) for v in valeurs]

        n = len(valeurs)
        if n == 1:
            presents[0][0][cle_cible][cle] = 50.0
        elif n > 1:
            for (t, _), v in zip(presents, valeurs):
                if sens == "H":
                    c = sum(1 for x in valeurs if x < v)
                else:
                    c = sum(1 for x in valeurs if x > v)
                t[cle_cible][cle] = c / (n - 1) * 100.0
        # les stocks sans valeur utilisable n'ont PAS de cle -> geres plus tard


def _piliers(t, cle_pct, poids_piliers):
    """Piliers + TOTAL + couverture. Chaque pilier = moyenne ponderee des
    metriques DISPONIBLES (poids renormalises). Un score None = donnee absente."""
    metr_par_pilier = {}
    for m in cfg.METRIQUES:
        metr_par_pilier.setdefault(m["pilier"], []).append(m)

    piliers = {}
    poids_dispo = 0.0   # poids effectif renseigne
    poids_tot = 0.0     # poids effectif total
    for pilier, metrs in metr_par_pilier.items():
        num = den = 0.0
        for mm in metrs:
            w = mm["poids"]
            poids_tot += w * poids_piliers[pilier]
            s = t[cle_pct].get(mm["cle"])
            if s is not None:
                num += s * w
                den += w
                poids_dispo += w * poids_piliers[pilier]
        piliers[pilier] = (num / den) if den > 0 else None

    piliers_ok = {p: v for p, v in piliers.items() if v is not None}
    sp = sum(poids_piliers[p] for p in piliers_ok)
    total = sum(piliers_ok[p] * poids_piliers[p] for p in piliers_ok) / sp if sp else None
    couverture = (poids_dispo / poids_tot) if poids_tot else 0.0
    return piliers, total, couverture


def _note(total):
    for lettre, seuil in cfg.GRILLE_NOTES:
        if total >= seuil:
            return lettre
    return cfg.GRILLE_NOTES[-1][0]


def _valuation(score_value):
    """Niveau de valorisation d'apres le score du pilier Value. Renvoie
    (libelle, index) ou index 0 = plus attractif."""
    for i, (libelle, seuil) in enumerate(cfg.NIVEAUX_VALUATION):
        if score_value >= seuil:
            return libelle, i
    return cfg.NIVEAUX_VALUATION[-1][0], len(cfg.NIVEAUX_VALUATION) - 1


def _forces_faiblesses(t):
    paires = [(cfg.NOMS_METRIQUES.get(c, c), p) for c, p in t["score_metrique"].items()
              if p is not None]
    forces = sorted((x for x in paires if x[1] >= cfg.SEUIL_FORCE),
                    key=lambda x: x[1], reverse=True)[:cfg.NB_FORCES]
    faibles = sorted((x for x in paires if x[1] <= cfg.SEUIL_FAIBLESSE),
                     key=lambda x: x[1])[:cfg.NB_FORCES]
    return forces, faibles


def _melange_scores(t):
    """score par metrique = MELANGE_RELATIF * percentile (secteur si dispo,
    sinon univers) + MELANGE_ABSOLU * score absolu (ancres).
    None = donnee absente (ne compte pas). Multiple de valo <= 0 = pire (0)."""
    t["score_metrique"] = {}
    src = t["pct_sect"] if (cfg.PERCENTILE_SECTORIEL and "pct_sect" in t) else t["pct"]
    for m in cfg.METRIQUES:
        cle = m["cle"]
        brut = t["brut"].get(cle)
        if brut is None:
            t["score_metrique"][cle] = None                 # absent -> renormalise
            continue
        if cle in cfg.NEGATIF_PIRE and brut <= 0:
            t["score_metrique"][cle] = 0.0                  # multiple negatif = pire
            continue
        rel = src.get(cle)
        if rel is None:
            t["score_metrique"][cle] = None
            continue
        absv = _score_absolu(cle, brut)
        if absv is None:
            t["score_metrique"][cle] = rel
        else:
            t["score_metrique"][cle] = (cfg.MELANGE_RELATIF * rel
                                        + cfg.MELANGE_ABSOLU * absv)


def _alertes(t):
    ops = {">": lambda a, b: a > b, ">=": lambda a, b: a >= b,
           "<": lambda a, b: a < b, "<=": lambda a, b: a <= b,
           "==": lambda a, b: a == b, "!=": lambda a, b: a != b}
    details = []
    for libelle, cle, op, seuil in cfg.REGLES_ALERTES:
        v = t["brut"].get(cle)
        if v is not None and ops[op](v, seuil):
            details.append(libelle)
    return details


def calculer_scores(titres, poids_piliers, winsoriser=True):
    n = len(titres)

    # 1) percentiles relatifs : univers complet, puis intra-secteur (>= SECTEUR_MIN)
    _percentiles_groupe(titres, "pct", winsoriser)
    secteurs = {}
    for t in titres:
        secteurs.setdefault(t["Secteur"], []).append(t)
    for secteur, groupe in secteurs.items():
        taille = len(groupe)
        for t in groupe:
            t["taille_secteur"] = taille
        if taille >= cfg.SECTEUR_MIN:
            _percentiles_groupe(groupe, "pct_sect", winsoriser)  # ajoute t["pct_sect"]

    # 2) score mixte par metrique -> piliers -> total, et derives
    for t in titres:
        _melange_scores(t)
        t["piliers"], t["total"], t["couverture"] = _piliers(t, "score_metrique", poids_piliers)
        assez = t["couverture"] >= getattr(cfg, "SEUIL_COUVERTURE", 0)
        t["note"] = _note(t["total"]) if (assez and t["total"] is not None) else "NR"
        det = _alertes(t)
        t["alertes"], t["alertes_detail"] = len(det), det
        t["conviction"] = (max(0.0, t["total"] - t["alertes"] * cfg.MALUS_ALERTE)
                           if t["total"] is not None else None)
        t["forces"], t["faiblesses"] = _forces_faiblesses(t)
        sv = t["piliers"].get("Value")
        if sv is None:
            t["valuation"], t["valo_niveau"] = "n/a", 99
        else:
            t["valuation"], t["valo_niveau"] = _valuation(sv)
        q, h = t["piliers"].get("Quality"), t["piliers"].get("Health")
        t["sweet_spot"] = (t["valo_niveau"] == 0 and assez
                           and q is not None and q >= cfg.SWEET_SPOT_QUALITE
                           and h is not None and h >= getattr(cfg, "SWEET_SPOT_SANTE", 0))

    # 2bis) au-dessus de la mediane de l'univers SUR Quality ET Value
    med_q = _mediane([t["piliers"].get("Quality") for t in titres])
    med_v = _mediane([t["piliers"].get("Value") for t in titres])
    for t in titres:
        q, v = t["piliers"].get("Quality"), t["piliers"].get("Value")
        t["qv_median"] = (med_q is not None and med_v is not None
                          and q is not None and q >= med_q
                          and v is not None and v >= med_v)

    # 3) classements univers (total, conviction) et rang intra-secteur
    #    (les None -- couverture nulle -- sont classes derniers)
    def _cle(x):
        return x if x is not None else -1.0
    ordre_total = sorted(range(n), key=lambda i: _cle(titres[i]["total"]), reverse=True)
    rangs_t = _rang_competition(ordre_total, lambda i: _cle(titres[i]["total"]))
    ordre_conv = sorted(range(n), key=lambda i: _cle(titres[i]["conviction"]), reverse=True)
    rangs_c = _rang_competition(ordre_conv, lambda i: _cle(titres[i]["conviction"]))
    for i, t in enumerate(titres):
        t["rang"], t["rang_conviction"] = rangs_t[i], rangs_c[i]
    for secteur, groupe in secteurs.items():
        ordre = sorted(range(len(groupe)), key=lambda k: _cle(groupe[k]["total"]), reverse=True)
        rangs_s = _rang_competition(ordre, lambda k: _cle(groupe[k]["total"]))
        for k, t in enumerate(groupe):
            t["rang_secteur"] = rangs_s[k]
    return titres


# ===========================================================================
#  Filtres
# ===========================================================================
def appliquer_filtres(titres, args):
    res = titres
    if args.min_score is not None:
        res = [t for t in res if t["total"] is not None and t["total"] >= args.min_score]
    if args.max_alertes is not None:
        res = [t for t in res if t["alertes"] <= args.max_alertes]
    if args.secteur:
        cibles = {_norm(s) for s in args.secteur}
        res = [t for t in res if _norm(t["Secteur"]) in cibles]
    if args.cap_min is not None:
        res = [t for t in res if (t["Cap"] or 0) >= args.cap_min]
    if args.note:
        notes_ok = {x.upper() for x in args.note}
        res = [t for t in res if t["note"].upper() in notes_ok]
    if args.valo_attractive:
        res = [t for t in res if t["valo_niveau"] == 0]
    if args.sweet_spot:
        res = [t for t in res if t["sweet_spot"]]
    for pilier, seuil in (args.pilier_min or {}).items():
        res = [t for t in res if (t["piliers"].get(pilier) or 0) >= seuil]
    cle_tri = "conviction" if args.classer_par == "conviction" else "total"
    res = sorted(res, key=lambda t: t[cle_tri] if t[cle_tri] is not None else -1.0, reverse=True)
    if args.top is not None:
        res = res[:args.top]
    return res


# ===========================================================================
#  Sorties console
# ===========================================================================
def _fmt_ff(paires):
    return ", ".join(f"{nom} ({p:.0f})" for nom, p in paires) if paires else "-"


def _f1(x):
    """Formatage 1 decimale tolerant au None."""
    return "n/a" if x is None else f"{x:.1f}"


def _rnd(x, n=1):
    """round() tolerant au None (renvoie None)."""
    return None if x is None else round(x, n)


def afficher_console(titres):
    entetes = ["Rg", "Ticker", "Secteur", "Qual", "Health", "Growth", "Value",
               "TOTAL", "Note", "Valuation", "SS", "RiskAdj", "Data", "Sect", "!", "Q&V"]
    lignes = []
    for t in titres:
        lignes.append([
            str(t["rang"]), t["Ticker"], t["Secteur"][:15],
            _f1(t["piliers"]["Quality"]), _f1(t["piliers"]["Health"]),
            _f1(t["piliers"]["Growth"]), _f1(t["piliers"]["Value"]),
            _f1(t["total"]), t["note"], t["valuation"],
            "*" if t["sweet_spot"] else "", _f1(t["conviction"]),
            f"{t['couverture']*100:.0f}",
            f"{t['rang_secteur']}/{t['taille_secteur']}", str(t["alertes"]),
            "X" if t.get("qv_median") else "",
        ])
    largeurs = [max(len(entetes[c]), *(len(l[c]) for l in lignes)) if lignes else len(entetes[c])
                for c in range(len(entetes))]
    gauche = {1, 2}

    def fmt(row):
        return "  ".join((row[c].ljust(largeurs[c]) if c in gauche else row[c].rjust(largeurs[c]))
                         for c in range(len(row)))
    print()
    print(fmt(entetes))
    print("  ".join("-" * largeurs[c] for c in range(len(entetes))))
    for l in lignes:
        print(fmt(l))
    print()


# ===========================================================================
#  Sorties CSV
# ===========================================================================
def ecrire_csv_resultats(titres, chemin):
    def r1(x):
        return round(x, 1) if x is not None else ""
    cols = ["Rang", "Ticker", "Secteur", "Cap ($Md)", "Quality", "Health", "Growth",
            "Value", "TOTAL", "Note", "Valuation", "Sweet spot", "Risk-adjusted score",
            "Rang risk-adj", "Data confidence (%)", "Rang secteur", "Taille secteur",
            "Alertes", "Detail alertes", "Q&V >= mediane", "Forces", "Faiblesses"]
    with open(chemin, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for t in sorted(titres, key=lambda x: x["rang"]):
            w.writerow([
                t["rang"], t["Ticker"], t["Secteur"],
                round(t["Cap"], 2) if t["Cap"] is not None else "",
                r1(t["piliers"]["Quality"]), r1(t["piliers"]["Health"]),
                r1(t["piliers"]["Growth"]), r1(t["piliers"]["Value"]),
                r1(t["total"]), t["note"], t["valuation"],
                "*" if t["sweet_spot"] else "", r1(t["conviction"]),
                t["rang_conviction"], round(t["couverture"] * 100),
                f"{t['rang_secteur']}/{t['taille_secteur']}",
                t["taille_secteur"], t["alertes"], " ; ".join(t["alertes_detail"]),
                "OK" if t.get("qv_median") else "",
                _fmt_ff(t["forces"]), _fmt_ff(t["faiblesses"]),
            ])


def ecrire_csv_scores(titres, chemin):
    cles = [m["cle"] for m in cfg.METRIQUES]
    with open(chemin, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Ticker"] + cles)
        for t in sorted(titres, key=lambda x: x["rang"]):
            w.writerow([t["Ticker"]] + [
                round(t["score_metrique"][c], 1) if t["score_metrique"].get(c) is not None else ""
                for c in cles])


# ===========================================================================
#  Sortie Excel (tableau de bord)
# ===========================================================================
def ecrire_excel(titres, tous_titres, poids, chemin, preset=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    BLEU = "1F3864"
    entete_font = Font(color="FFFFFF", bold=True)
    entete_fill = PatternFill("solid", fgColor=BLEU)
    titre_font = Font(color=BLEU, bold=True, size=14)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bord = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    rouge = PatternFill("solid", fgColor="FFC7CE")
    echelle = lambda: ColorScaleRule(
        start_type="num", start_value=0, start_color="F8696B",
        mid_type="num", mid_value=50, mid_color="FFEB84",
        end_type="num", end_value=100, end_color="63BE7B")

    def style_entete(ws, ncol, ligne=1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=ligne, column=c)
            cell.fill, cell.font, cell.alignment = entete_fill, entete_font, centre

    titres_tri = sorted(titres, key=lambda x: x["rang"])
    wb = Workbook()

    # ---- 1) SYNTHESE ------------------------------------------------------
    ws = wb.active
    ws.title = "Synthese"
    ws["A1"] = "QS SCREENER - Synthese"
    ws["A1"].font = titre_font
    ws["A2"] = (f"Univers : {len(tous_titres)} titres  |  affiches : {len(titres_tri)}  |  "
                f"poids " + " / ".join(f"{p} {poids[p]}" for p in PILIERS)
                + (f"  |  preset {preset}" if preset else ""))
    ws["A3"] = "Note = score TOTAL (A+..D). Conviction = TOTAL ajuste du risque (malus par alerte)."

    r = 5
    ws.cell(r, 1, "TOP CLASSEMENT").font = Font(bold=True, color=BLEU)
    r += 1
    cols = ["Rang", "Ticker", "Secteur", "Cap ($Md)", "Quality", "Health",
            "Growth", "Value", "TOTAL", "Note", "Conviction", "Alertes"]
    for c, h in enumerate(cols, 1):
        ws.cell(r, c, h)
    style_entete(ws, len(cols), r)
    debut = r + 1
    for t in titres_tri:
        r += 1
        vals = [t["rang"], t["Ticker"], t["Secteur"],
                round(t["Cap"], 1) if t["Cap"] is not None else None,
                _rnd(t["piliers"]["Quality"]), _rnd(t["piliers"]["Health"]),
                _rnd(t["piliers"]["Growth"]), _rnd(t["piliers"]["Value"]),
                _rnd(t["total"]), t["note"], _rnd(t["conviction"]), t["alertes"]]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)
        if t["alertes"]:
            ws.cell(r, 12).fill = rouge
    for col in range(5, 12):  # Quality..Conviction
        L = get_column_letter(col)
        ws.conditional_formatting.add(f"{L}{debut}:{L}{r}", echelle())

    # tableau par secteur
    r += 2
    ws.cell(r, 1, "MOYENNES PAR SECTEUR").font = Font(bold=True, color=BLEU)
    r += 1
    sec_cols = ["Secteur", "Nb", "TOTAL moy", "Quality", "Health", "Growth", "Value"]
    for c, h in enumerate(sec_cols, 1):
        ws.cell(r, c, h)
    style_entete(ws, len(sec_cols), r)
    secteurs = {}
    for t in tous_titres:
        secteurs.setdefault(t["Secteur"], []).append(t)
    debut_sec = r + 1
    for secteur, grp in sorted(secteurs.items(),
                               key=lambda kv: sum(x["total"] for x in kv[1]) / len(kv[1]),
                               reverse=True):
        r += 1
        moy = lambda k: sum(x["piliers"][k] for x in grp) / len(grp)
        vals = [secteur, len(grp), round(sum(x["total"] for x in grp) / len(grp), 1),
                round(moy("Quality"), 1), round(moy("Health"), 1),
                round(moy("Growth"), 1), round(moy("Value"), 1)]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)
    for col in range(3, 8):
        L = get_column_letter(col)
        ws.conditional_formatting.add(f"{L}{debut_sec}:{L}{r}", echelle())

    for i, wdt in enumerate([6, 9, 16, 10, 9, 9, 9, 9, 9, 6, 11, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A6"

    # ---- 2) RESULTATS -----------------------------------------------------
    ws2 = wb.create_sheet("Resultats")
    cols2 = ["Rang", "Ticker", "Secteur", "Cap ($Md)", "Quality", "Health", "Growth",
             "Value", "TOTAL", "Note", "Conviction", "Rg conv.", "Rg secteur",
             "Alertes", "Detail alertes", "Forces", "Faiblesses"]
    ws2.append(cols2)
    for t in titres_tri:
        ws2.append([
            t["rang"], t["Ticker"], t["Secteur"],
            round(t["Cap"], 1) if t["Cap"] is not None else None,
            _rnd(t["piliers"]["Quality"]), _rnd(t["piliers"]["Health"]),
            _rnd(t["piliers"]["Growth"]), _rnd(t["piliers"]["Value"]),
            _rnd(t["total"]), t["note"], _rnd(t["conviction"]),
            t["rang_conviction"], f"{t['rang_secteur']}/{t['taille_secteur']}",
            t["alertes"], " ; ".join(t["alertes_detail"]),
            _fmt_ff(t["forces"]), _fmt_ff(t["faiblesses"]),
        ])
    style_entete(ws2, len(cols2))
    ws2.freeze_panes = "A2"
    nrow = len(titres_tri) + 1
    for col in range(5, 12):
        L = get_column_letter(col)
        ws2.conditional_formatting.add(f"{L}2:{L}{nrow}", echelle())
    for rr in range(2, nrow + 1):
        if ws2.cell(rr, 14).value:
            ws2.cell(rr, 14).fill = rouge
    for i, wdt in enumerate([6, 9, 16, 10, 9, 9, 9, 9, 9, 6, 11, 8, 10, 8, 34, 40, 40], 1):
        ws2.column_dimensions[get_column_letter(i)].width = wdt

    # ---- 3) SCORES (percentiles) -----------------------------------------
    ws3 = wb.create_sheet("Scores")
    cles = [m["cle"] for m in cfg.METRIQUES]
    ws3.append(["Ticker"] + cles)
    for t in titres_tri:
        ws3.append([t["Ticker"]] + [(_rnd(t["score_metrique"].get(c))) for c in cles])
    style_entete(ws3, len(cles) + 1)
    ws3.freeze_panes = "B2"
    nrow3 = len(titres_tri) + 1
    for col in range(2, len(cles) + 2):
        L = get_column_letter(col)
        ws3.conditional_formatting.add(f"{L}2:{L}{nrow3}", echelle())
    ws3.column_dimensions["A"].width = 9
    for col in range(2, len(cles) + 2):
        ws3.column_dimensions[get_column_letter(col)].width = 8

    # ---- 4) DATA (valeurs brutes) ----------------------------------------
    ws4 = wb.create_sheet("Data")
    ref_cles = [r["cle"] for r in cfg.COLONNES_REFERENCE]
    ws4.append(["Ticker", "Secteur", "Cap ($Md)"] + cles + ref_cles)
    for t in titres_tri:
        ligne = [t["Ticker"], t["Secteur"],
                 round(t["Cap"], 2) if t["Cap"] is not None else None]
        ligne += [t["brut"].get(c) for c in cles]
        ligne += [t["ref"].get(c) for c in ref_cles]
        ws4.append(ligne)
    style_entete(ws4, 3 + len(cles) + len(ref_cles))
    ws4.freeze_panes = "B2"
    ws4.column_dimensions["A"].width = 9
    ws4.column_dimensions["B"].width = 16

    # ---- 5) FICHES (une par action) --------------------------------------
    ws5 = wb.create_sheet("Fiches")
    rr = 1
    for t in titres_tri:
        ws5.cell(rr, 1, f"#{t['rang']}  {t['Ticker']}  -  {t['Secteur']}").font = \
            Font(bold=True, size=12, color=BLEU)
        rr += 1
        ws5.cell(rr, 1, f"TOTAL {t['total']:.1f} (Note {t['note']})   "
                        f"Conviction {t['conviction']:.1f}   Alertes : {t['alertes']}")
        rr += 1
        ws5.cell(rr, 1, f"Quality {t['piliers']['Quality']:.0f} | "
                        f"Health {t['piliers']['Health']:.0f} | "
                        f"Growth {t['piliers']['Growth']:.0f} | "
                        f"Value {t['piliers']['Value']:.0f}")
        rr += 1
        ws5.cell(rr, 1, "Forces : " + _fmt_ff(t["forces"]))
        rr += 1
        ws5.cell(rr, 1, "Faiblesses : " + _fmt_ff(t["faiblesses"]))
        rr += 1
        if t["alertes_detail"]:
            c = ws5.cell(rr, 1, "Risques : " + " ; ".join(t["alertes_detail"]))
            c.font = Font(color="C00000")
            rr += 1
        rr += 1
    ws5.column_dimensions["A"].width = 100

    # ---- 6) CONFIG (rappel des reglages) ---------------------------------
    ws6 = wb.create_sheet("Config")
    ws6.append(["Poids des piliers"])
    ws6.append(PILIERS)
    ws6.append([poids[p] for p in PILIERS])
    ws6.append([])
    ws6.append(["Metrique", "Pilier", "Poids intra", "Sens (H=haut mieux)"])
    for m in cfg.METRIQUES:
        ws6.append([cfg.NOMS_METRIQUES.get(m["cle"], m["cle"]), m["pilier"],
                    m["poids"], m["sens"]])
    ws6.append([])
    ws6.append(["Regles d'alertes"])
    for lib, cle, op, seuil in cfg.REGLES_ALERTES:
        ws6.append([lib, f"{cle} {op} {seuil}"])
    ws6.column_dimensions["A"].width = 26
    ws6.column_dimensions["B"].width = 22

    wb.save(chemin)
    return True


# ===========================================================================
#  CLI
# ===========================================================================
def _parse_pilier_min(valeurs):
    d = {}
    for v in valeurs or []:
        if "=" not in v:
            raise SystemExit(f"--pilier-min attend pilier=valeur (recu: {v})")
        cle, num = v.split("=", 1)
        cle = cle.strip().capitalize()
        if cle not in PILIERS:
            raise SystemExit(f"Pilier inconnu: {cle}. Choix: {', '.join(PILIERS)}")
        d[cle] = float(num)
    return d


def main(argv=None):
    p = argparse.ArgumentParser(
        description="QS Screener v3 - screener d'actions Quality (systeme QS v2 etendu).",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("csv", help="Fichier CSV contenant les donnees de l'univers.")
    p.add_argument("--preset", choices=list(cfg.PRESETS.keys()), default=None,
                   help="Preset de poids des piliers.")
    p.add_argument("--classer-par", choices=["total", "conviction"], default="total",
                   help="Critere de tri/classement affiche (defaut: total).")
    p.add_argument("--min-score", type=float, help="Ne garder que TOTAL >= N.")
    p.add_argument("--max-alertes", type=int, help="Ne garder que Alertes <= N.")
    p.add_argument("--note", action="append", help="Filtrer une note-lettre (ex: A+, repetable).")
    p.add_argument("--valo-attractive", action="store_true",
                   help="Ne garder que les valos 'Attractive'.")
    p.add_argument("--sweet-spot", action="store_true",
                   help="Ne garder que la cible du projet : qualite solide + valo attractive.")
    p.add_argument("--secteur", action="append", help="Filtrer un secteur (repetable).")
    p.add_argument("--cap-min", type=float, help="Capitalisation minimale ($Md).")
    p.add_argument("--pilier-min", action="append",
                   help="Seuil minimal sur un pilier, ex: Quality=60 (repetable).")
    p.add_argument("--top", type=int, help="Ne garder que les N premiers.")
    p.add_argument("--sans-winsor", action="store_true", help="Desactive la winsorisation.")
    p.add_argument("--delimiter", help="Force le separateur CSV (ex: ';').")
    p.add_argument("--out-dir", default="qs_out", help="Dossier de sortie (defaut: qs_out).")
    p.add_argument("--no-images", action="store_true", help="N'ecrit pas les PNG.")
    p.add_argument("--pdf", action="store_true", help="Ecrit aussi le rapport en PDF.")
    p.add_argument("--excel", action="store_true", help="Ecrit aussi un classeur Excel.")
    p.add_argument("--dpi", type=int, default=200, help="Resolution des PNG (defaut: 200).")
    p.add_argument("--no-fichiers", action="store_true", help="Console uniquement.")
    p.add_argument("--open", action="store_true", dest="ouvrir",
                   help="Ouvre l'image dashboard a la fin (macOS).")
    args = p.parse_args(argv)
    args.pilier_min = _parse_pilier_min(args.pilier_min)

    poids = dict(cfg.PRESETS[args.preset]) if args.preset else dict(cfg.POIDS_PILIERS)

    titres, manquantes = charger_csv(args.csv, delimiteur=args.delimiter)
    if manquantes:
        noms = ", ".join(cfg.NOMS_METRIQUES.get(m, m) for m in manquantes)
        print(f"[!] Metriques absentes du CSV (neutralisees a 50) : {noms}", file=sys.stderr)

    calculer_scores(titres, poids, winsoriser=not args.sans_winsor)
    retenus = appliquer_filtres(titres, args)

    print(f"\nQS Screener v3  |  univers = {len(titres)} titres  |  "
          + " / ".join(f"{k} {poids[k]}" for k in PILIERS)
          + (f"  |  preset '{args.preset}'" if args.preset else "")
          + (f"  |  tri: {args.classer_par}" if args.classer_par != "total" else ""))
    if len(retenus) != len(titres):
        print(f"Filtres actifs -> {len(retenus)} titre(s) retenu(s).")
    afficher_console(retenus)

    if not args.no_fichiers:
        os.makedirs(args.out_dir, exist_ok=True)
        chemin_res = os.path.join(args.out_dir, "resultats.csv")
        chemin_sco = os.path.join(args.out_dir, "scores_detail.csv")
        ecrire_csv_resultats(retenus, chemin_res)
        ecrire_csv_scores(retenus, chemin_sco)
        produits = [chemin_res, chemin_sco]
        a_ouvrir = []

        if not args.no_images:
            images = qs_pdf.generer_images(retenus, titres, poids, args.out_dir,
                                           preset=args.preset, dpi=args.dpi)
            if images:
                produits.extend(images)
                a_ouvrir = list(images)  # ouvre les 2 pages (dashboard + methodology)
            else:
                print("[i] Images PNG ignorees : il manque fpdf2 et/ou pymupdf. "
                      "Installe-les avec : ./.venv/bin/pip install fpdf2 pymupdf",
                      file=sys.stderr)

        if args.pdf:
            chemin_pdf = os.path.join(args.out_dir, "QS_Screener_rapport.pdf")
            if qs_pdf.generer_pdf(retenus, titres, poids, chemin_pdf, preset=args.preset):
                produits.append(chemin_pdf)
                if not a_ouvrir:
                    a_ouvrir = [chemin_pdf]
            else:
                print("[i] fpdf2 absent : PDF ignore.", file=sys.stderr)

        if args.excel:
            chemin_xlsx = os.path.join(args.out_dir, "QS_Screener_resultats.xlsx")
            if ecrire_excel(retenus, titres, poids, chemin_xlsx, preset=args.preset):
                produits.append(chemin_xlsx)
            else:
                print("[i] openpyxl absent : export Excel ignore.", file=sys.stderr)

        print("Fichiers ecrits :")
        for c in produits:
            print("   -", c)
        print()
        if args.ouvrir and a_ouvrir and sys.platform == "darwin":
            subprocess.run(["open", *a_ouvrir], check=False)


if __name__ == "__main__":
    main()
